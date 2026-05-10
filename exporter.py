from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _autofit(ws):
    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 3, 45)


def _style_header(ws):
    fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def export_to_excel(df: pd.DataFrame, buckets: dict) -> BytesIO:
    """
    Build a two-sheet workbook:
      • Transactions  — all rows with date, description, amount, category, sub_category, bucket
      • Summary       — total spend per category and bucket
    """
    output = BytesIO()

    # Add bucket column derived from category → buckets config
    export_df = df.copy()
    export_df["bucket"] = export_df["category"].map(buckets).fillna("Others")

    # Reorder columns for the output sheet
    cols = ["date", "description", "amount", "category", "sub_category", "bucket"]
    # Keep any extra columns (debit/credit) at the end
    extra = [c for c in export_df.columns if c not in cols]
    export_df = export_df[cols + extra]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # --- Sheet 1: Transactions ---
        export_df.to_excel(writer, sheet_name="Transactions", index=False)
        _style_header(writer.sheets["Transactions"])
        _autofit(writer.sheets["Transactions"])

        # --- Sheet 2: Summary by category ---
        summary = (
            export_df.groupby(["bucket", "category"])["amount"]
            .sum()
            .reset_index()
            .rename(columns={"amount": "Total (INR)"})
            .sort_values(["bucket", "Total (INR)"], ascending=[True, False])
        )
        summary.to_excel(writer, sheet_name="Summary", index=False)
        _style_header(writer.sheets["Summary"])
        _autofit(writer.sheets["Summary"])

    output.seek(0)
    return output
